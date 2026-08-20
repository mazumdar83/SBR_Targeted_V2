"""Build the run workbook from stage outputs. Skips sheets whose inputs are absent."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HDR = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
HFILL = PatternFill("solid", fgColor="1F3864")
BODY = Font(name=ARIAL, size=10)
FILLS = {
    "CURATED": PatternFill("solid", fgColor="FFF2CC"),
    "AMBIGUOUS": PatternFill("solid", fgColor="FCE4D6"),
    "ANCHOR_NEW_ON_TERM": PatternFill("solid", fgColor="C6E0B4"),
    "ANCHOR_ADJACENT_OFF_TERM": PatternFill("solid", fgColor="FCE4D6"),
    "NOT_IN_KO": PatternFill("solid", fgColor="F2F2F2"),
    "OFF_TERM": PatternFill("solid", fgColor="F2F2F2"),
}

SHEETS = [
    ("gene_ko_map.csv", "Gene_KO_Map", "status_code"),
    ("kegg_anchor_table.tsv", "KEGG_Anchors", "anchor_class"),
    ("discovered_kos.csv", "Discovered_KOs", None),
    ("ko_coverage_gaps.csv", "KO_Coverage_Gaps", None),
    ("taxonomy_mapping.tsv", "Taxonomy_Mapping", None),
    ("rejected_symbol_collisions.tsv", "Rejected_Collisions", None),
    ("rejected_ko_collisions.csv", "Rejected_KO_Collisions", None),
    ("discovered_pfam.csv", "Discovered_Pfam", None),
    ("ko_sequence_manifest.csv", "KO_Sequences", None),
    ("ko_still_uncovered.csv", "KO_Still_Uncovered", None),
]


def _sheet(ws, df: pd.DataFrame, colour_col=None):
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(1, c)
        cell.font, cell.fill = HDR, HFILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = 26
    ws.row_dimensions[1].height = 28
    for r in df.itertuples(index=False):
        ws.append([str(v)[:32000] if pd.notna(v) else "" for v in r])
    idx = list(df.columns).index(colour_col) if colour_col in df.columns else None
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = BODY
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if idx is not None:
            v = str(df.iloc[i - 2, idx])
            if v in FILLS:
                for cell in row:
                    cell.fill = FILLS[v]
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{ws.max_row}"


def build(run: Path) -> Path:
    run = Path(run)
    wb = Workbook()
    wb.remove(wb.active)
    prot = run / "proteins_anchored.tsv"
    if not prot.exists():
        prot = run / "proteins.tsv"
    if prot.exists():
        df = pd.read_csv(prot, sep="\t", low_memory=False)
        _sheet(wb.create_sheet("Protein_Manifest"),
               df.drop(columns=[c for c in ["Sequence"] if c in df.columns]))
    for fname, sheet, colour in SHEETS:
        p = run / fname
        if not p.exists():
            continue
        sep = "\t" if p.suffix == ".tsv" else ","
        try:
            d = pd.read_csv(p, sep=sep, low_memory=False)
        except pd.errors.EmptyDataError:
            continue          # stage wrote a header-less empty file; nothing to report
        if d.empty:
            continue
        _sheet(wb.create_sheet(sheet), d, colour)

    ws = wb.create_sheet("Method")
    lex = run / "term_lexicon.expanded.json"
    if not lex.exists():
        lex = run / "term_lexicon.json"
    import json
    L = json.loads(lex.read_text()) if lex.exists() else {}
    lines = [
        (f"bfgm run: {L.get('term','(unknown term)')}", True), ("", False),
        ("PIPELINE", True),
        ("stage 0 seed        literature to genes, audited, critic-reviewed", False),
        ("stage 1 ko-map      gene symbols to KEGG KOs, exact token match", False),
        ("stage 2 uniprot     metadata only, two axes: gene symbol and discovered Pfam", False),
        ("stage 3 sequences   sequences and taxonomic lineage for the curated set", False),
        ("stage 4 kegg-anchor accession to KEGG gene to KO, validates the whole chain", False),
        ("", False),
        ("KO IS NOT A UNIPROT JOIN KEY", True),
        ("UniProt carries no KEGG Orthology cross-reference. Verified on P06971: the entry", False),
        ("has KEGG gene IDs, eggNOG, InterPro and Pfam, but no KO line, and xref:ko- returns", False),
        ("zero. Stage 4 therefore builds the link in two hops through xref_kegg and the", False),
        ("per-organism /link/ko/<org> endpoint. The global /link/genes/ko is unreliable.", False),
        ("", False),
        ("SYMBOL COLLISIONS ARE THE MAIN FAILURE MODE", True),
        ("Gene symbols collide heavily across taxa. Both stage 1 and stage 2 resolve", False),
        ("collisions against the term lexicon rather than by hardcoded rules, and every", False),
        ("rejection is written to a Rejected sheet with the definition that triggered it.", False),
        ("", False),
        ("LEXICON", True),
        (f"synonyms: {', '.join(L.get('synonyms', []))[:400]}", False),
        (f"negative terms: {', '.join(L.get('negative_terms', []))[:400]}", False),
        ("", False),
        ("LICENSING", True),
        ("KO columns make this a KEGG-derived asset; commercial use needs a Pathway", False),
        ("Solutions licence. UniProt sequences and taxonomy are CC BY 4.0 and carry no", False),
        ("such restriction, so dropping the KO columns leaves a freely usable asset.", False),
    ]
    for i, (t, b) in enumerate(lines, 1):
        ws.cell(i, 1, t).font = Font(name=ARIAL, size=10, bold=b)
    ws.column_dimensions["A"].width = 110

    outp = run / f"{run.name}_report.xlsx"
    wb.save(outp)
    return outp
